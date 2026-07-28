from io import BytesIO
import csv
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.response import Response
from rest_framework.views import APIView
from ..pagination import CustomPagination
from openpyxl import Workbook
from openpyxl.styles import (PatternFill,Font,Alignment,Border,Side,)
from reportlab.platypus import (SimpleDocTemplate,Table,Spacer,)
from pos.models import (Product,StockMovement,StockAdjustment,StockAdjustmentItem,Company,)
from rest_framework import status
from ..services.inventory import *
from ..filters import *
from ..serializers import *
from ..pdf_helpers import *
from ..helpers import *
from ..Order_no_generator import *
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta




class TestInventoryView(APIView):

    def post(self, request):

        product = Product.objects.get(id=1)

        update_stock(
            product=product,
            quantity=-2,
            movement_type="sale",
            reference="TEST-001",
            note="Inventory engine test",
        )

        return Response({
            "message": "Stock updated successfully."
        })
    

class StockAdjustmentView(APIView):

    @transaction.atomic
    def post(self, request):

        serializer = StockAdjustmentSerializer(
            data=request.data
        )

        serializer.is_valid(
            raise_exception=True
        )

        data = serializer.validated_data

        product = get_object_or_404(
            Product,
            id=data["product_id"]
        )

        quantity = data["quantity"]

        movement_type = data["movement_type"]

        note = data.get("note", "")

        if movement_type == "damage":
            quantity = -abs(quantity)

        elif movement_type == "purchase":
            quantity = abs(quantity)

        elif movement_type == "adjustment":
            pass 

        new_stock = product.stock_quantity + quantity

        if new_stock < 0:
        
            return Response(
                {
                    "error": "Stock cannot become negative."
                },
                status=status.HTTP_400_BAD_REQUEST
            )  

        update_stock(
            product=product,
            quantity=quantity,
            movement_type=movement_type,
            reference="MANUAL",
            note=note,
        )                 

        return Response({
            "success": True,
            "product": product.name,
            "current_stock": product.stock_quantity,
        }) 


class InventoryListView(APIView):

    def get(self, request):

        print(request.GET)

        products = get_filtered_inventory(request)

        search = request.GET.get("search")

        if search:
        
            products = products.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(barcode__icontains=search)
            )

        category = request.GET.get("category")

        if category:
        
            products = products.filter(
                category_id=category
            )

        active = request.GET.get("active")

        if active == "true":
            products = products.filter(
                is_active=True
            )

        elif active == "false":
            products = products.filter(
                is_active=False
            )

        status = request.GET.get("status")

        if status == "out":
        
            products = products.filter(
                stock_quantity=0
            )

        elif status == "low":
        
            products = products.filter(
                stock_quantity__gt=0,
                stock_quantity__lte=F("min_stock")
            )

        elif status == "inactive":
        
            products = products.filter(
                is_active=False
            )

        sort = request.GET.get("sort", "name")

        allowed = [
        
            "name",

            "-name",

            "stock_quantity",

            "-stock_quantity",

            "sales_price",

            "-sales_price",

            "cost_price",

            "-cost_price",

        ]

        if sort in allowed:
        
            products = products.order_by(sort)        

        paginator = CustomPagination()

        page = paginator.paginate_queryset(
            products,
            request
        )

        serializer = InventorySerializer(
            page,
            many=True
        )

        return paginator.get_paginated_response(
            serializer.data
        )  


class StockMovementHistoryView(APIView):

    def get(self, request):

        movements = get_filtered_stock_movements(request)

        paginator = CustomPagination()

        page = paginator.paginate_queryset(
            movements,
            request
        )

        serializer = StockMovementSerializer(
            page,
            many=True
        )

        return paginator.get_paginated_response(
            serializer.data
        )              
    

class StockMovementSummaryView(APIView):

    def get(self, request):

        movements = StockMovement.objects.all()

        today = timezone.now().date()

        today_start = timezone.make_aware(
            timezone.datetime.combine(
                today,
                timezone.datetime.min.time()
            )
        )

        week_start = today_start - timedelta(days=today.weekday())

        month_start = today_start.replace(day=1)

        last30_start = today_start - timedelta(days=30)

        total_movements = movements.count()

        today_movements = movements.filter(
            created_at__gte=today_start
        ).count()

        week_movements = movements.filter(
            created_at__gte=week_start
        ).count()

        month_movements = movements.filter(
            created_at__gte=month_start
        ).count()

        last30_movements = movements.filter(
            created_at__gte=last30_start
        ).count()

        today_stock_added = (
            movements.filter(
                created_at__gte=today_start,
                quantity__gt=0
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        today_stock_removed = abs(
            movements.filter(
                created_at__gte=today_start,
                quantity__lt=0
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )
        
        week_stock_added = (
            movements.filter(
                created_at__gte=week_start,
                quantity__gt=0
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        week_stock_removed = abs(
            movements.filter(
                created_at__gte=week_start,
                quantity__lt=0
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        month_stock_added = (
            movements.filter(
                created_at__gte=month_start,
                quantity__gt=0
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        month_stock_removed = abs(
            movements.filter(
                created_at__gte=month_start,
                quantity__lt=0
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        stock_added = (
            movements
            .filter(quantity__gt=0)
            .aggregate(total=Sum("quantity"))["total"] or 0
        )

        stock_removed = abs(
            movements
            .filter(quantity__lt=0)
            .aggregate(total=Sum("quantity"))["total"] or 0
        )

        sales_count = movements.filter(
            movement_type="sale"
        ).count()

        sales_quantity = abs(
            movements.filter(
                movement_type="sale"
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        purchase_count = movements.filter(
            movement_type="purchase"
        ).count()

        purchase_quantity = (
            movements.filter(
                movement_type="purchase"
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        refund_count = movements.filter(
            movement_type="refund"
        ).count()

        refund_quantity = (
            movements.filter(
                movement_type="refund"
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        adjustment_count = movements.filter(
            movement_type="adjustment"
        ).count()

        adjustment_quantity = (
            movements.filter(
                movement_type="adjustment"
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        damage_count = movements.filter(
            movement_type="damage"
        ).count()

        damage_quantity = abs(
            movements.filter(
                movement_type="damage"
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        purchase_return_count = movements.filter(
            movement_type="purchase_return"
        ).count()

        purchase_return_quantity = abs(
            movements.filter(
                movement_type="purchase_return"
            ).aggregate(total=Sum("quantity"))["total"] or 0
        )

        net_change = stock_added - stock_removed

        daily_chart = (
            StockMovement.objects
            .filter(created_at__gte=last30_start)
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(
            
                stock_in=Sum(
                    "quantity",
                    filter=Q(quantity__gt=0)
                ),

                stock_out=Sum(
                    "quantity",
                    filter=Q(quantity__lt=0)
                ),

                movements=Count("id")

            )
            .order_by("day")
        )

        chart_data = []

        for row in daily_chart:
        
            chart_data.append({
            
                "date": row["day"],

                "movements": row["movements"],

                "stock_in": row["stock_in"] or 0,

                "stock_out": abs(row["stock_out"] or 0),

            })

        return Response({

            "total_movements": total_movements,

            "stock_added": stock_added,

            "stock_removed": stock_removed,

            "net_change": net_change,

            "sales": {
                "count": sales_count,
                "quantity": sales_quantity,
            },

            "purchases": {
                "count": purchase_count,
                "quantity": purchase_quantity,
            },

            "refunds": {
                "count": refund_count,
                "quantity": refund_quantity,
            },

            "adjustments": {
                "count": adjustment_count,
                "quantity": adjustment_quantity,
            },

            "damages": {
                "count": damage_count,
                "quantity": damage_quantity,
            },

            "purchase_returns": {
                "count": purchase_return_count,
                "quantity": purchase_return_quantity,
            },

            "time_periods": {

            "today": {
            
                "movements": today_movements,

                "stock_added": today_stock_added,

                "stock_removed": today_stock_removed,

            },

            "this_week": {
            
                "movements": week_movements,

                "stock_added": week_stock_added,

                "stock_removed": week_stock_removed,

            },

            "this_month": {
            
                "movements": month_movements,

                "stock_added": month_stock_added,

                "stock_removed": month_stock_removed,

            },

            "last_30_days": {
            
                "movements": last30_movements,

            }
            },
            "charts": {

            "daily_movements": chart_data

        },



        

        })    
    

class StockMovementExportCSVView(APIView):

    def get(self, request):

        movements = get_filtered_stock_movements(request)

        response = HttpResponse(
            content_type="text/csv"
        )

        filename = build_export_filename(
            request,
            "stock-movements",
            "csv"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        

        writer = csv.writer(response)

        writer.writerow([
            "S.No",
            "Product",
            "Movement",
            "Quantity",
            "Previous Stock",
            "New Stock",
            "Reference",
            "Date",
        ])

        for index, movement in enumerate(movements, start=1):

            if movement.quantity > 0:
                quantity = f"+{movement.quantity}"
            else:
                quantity = str(movement.quantity)

            writer.writerow([

                index,

                movement.product.name,

                movement.get_movement_type_display(),

                quantity,

                movement.previous_stock,

                movement.new_stock,

                movement.reference,

                movement.created_at.strftime("%d-%b-%Y %I:%M %p"),

            ])

        return response    
    

class StockMovementExportExcelView(APIView):

    def get(self, request):

        movements = get_filtered_stock_movements(request)

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Stock Movements"

        headers = [
        
            "S.No",

            "Product",

            "Movement",

            "Quantity",

            "Previous Stock",

            "New Stock",

            "Reference",

            "Date",

        ]

        sheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="2E7D32"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        thin_border = Border(
        
            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin"),
        )

        for cell in sheet[1]:
        
            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = header_alignment

            cell.border = thin_border

        for index, movement in enumerate(movements, start=1):
        
            quantity = (
                f"+{movement.quantity}"
                if movement.quantity > 0
                else str(movement.quantity)
            )

            sheet.append([
            
                index,

                movement.product.name,

                movement.get_movement_type_display(),

                quantity,

                movement.previous_stock,

                movement.new_stock,

                movement.reference,

                movement.created_at.strftime("%d-%b-%Y %I:%M %p"),

            ])

        for row in sheet.iter_rows(
            min_row=2
        ):

            for cell in row:
            
                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="center"
                )


        for column in sheet.columns:
        
            length = max(
                len(str(cell.value or ""))
                for cell in column
            )

            sheet.column_dimensions[
                column[0].column_letter
            ].width = length + 3

        sheet.freeze_panes = "A2"

        response = HttpResponse(
        
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        filename = build_export_filename(
            request,
            "stock-movements",
            "xlsx"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        workbook.save(response)

        return response
    

class StockMovementExportPDFView(APIView):

    def get(self, request):

        buffer = BytesIO()

        doc = SimpleDocTemplate(buffer)

        styles = getSampleStyleSheet()

        story = []

        company = Company.objects.first()

        movements = get_filtered_stock_movements(request)    

        add_company_header(
            story,
            company,
        )
        
        add_report_title(
            story,
            "STOCK MOVEMENT REPORT"
        )
        
        story.append(
            Spacer(1, 15)
        )
        
        report_info = [
        
            ["Generated", timezone.localtime().strftime("%d-%b-%Y %I:%M %p")]
        
        ]

        movement = request.GET.get("movement")

        product = request.GET.get("product")

        category = request.GET.get("category")

        date = request.GET.get("date")

        if movement:
            report_info.append(["Movement", movement])

        if product:
            report_info.append(["Product", product])

        if category:
            report_info.append(["Category", category])

        if date:
            report_info.append(["Date Filter", date.title()])

        add_report_info(
            story,
            report_info
        )            

        table_data = [[
        
            "S.No",

            "Product",

            "Movement",

            "Qty",

            "Previous",

            "New",

            "Reference",

            "Date",

        ]]

        for index, movement in enumerate(movements, start=1):
        
            qty = (
                f"+{movement.quantity}"
                if movement.quantity > 0
                else str(movement.quantity)
            )

            table_data.append([
            
                index,

                movement.product.name,

                movement.get_movement_type_display(),

                qty,

                movement.previous_stock,

                movement.new_stock,

                movement.reference,

                movement.created_at.strftime("%d-%b-%Y"),

            ])

        table = Table(
        
            table_data,

            colWidths=[35, 120, 70, 45, 55, 55, 80, 70],

            repeatRows=1

        )

        style_report_table(
            table,
            len(table_data),
            "#2E7D32"
        )

        story.append(table)

        total = movements.count()

        stock_added = sum(
            m.quantity
            for m in movements
            if m.quantity > 0
        )

        stock_removed = abs(sum(
            m.quantity
            for m in movements
            if m.quantity < 0
        ))

        net_change = stock_added - stock_removed        

        summary = [
        
            ["Total Movements", str(total)],

            ["Stock Added", str(stock_added)],

            ["Stock Removed", str(stock_removed)],

            ["Net Change", str(net_change)]

        ]

        add_summary_table(
            story,
            "SUMMARY",
            summary
        )
        doc.build(

            story,

            onFirstPage=add_page_footer,

            onLaterPages=add_page_footer,

        )

        pdf = buffer.getvalue()
        
        buffer.close()
        
        response = HttpResponse(
            pdf,
            content_type="application/pdf"
        )
        
        filename = build_export_filename(
            request,
            "stock-movements",
            "pdf"
        )
        
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        
        return response   


class InventorySummaryView(APIView):

    def get(self, request):

        data = get_inventory_summary()

        return Response(data)


class LowStockProductsView(APIView):

    def get(self, request):

        products = get_filtered_low_stock_products(
            request
        )

        serializer = ProductSerializer(
            products,
            many=True
        )

        return Response(serializer.data)


class ProductDetailView(APIView):

    def get(self, request, pk):

        product = get_object_or_404(Product, pk=pk)

        product_data = ProductDetailSerializer(product).data

        recent_movements = StockMovement.objects.filter(
            product=product
        ).order_by("-created_at")[:10]

        movement_data = StockMovementSerializer(
            recent_movements,
            many=True,
        ).data

        return Response({"product": product_data,"recent_movements": movement_data,})
    

class InventoryDashboardView(APIView):

    def get(self, request):

        data = get_inventory_statistics()

        return Response(data)


class StockAdjustmentView(APIView):

    @transaction.atomic
    def post(self, request):

        serializer = StockAdjustmentSerializer(
            data=request.data
        )

        serializer.is_valid(
            raise_exception=True
        )

        data = serializer.validated_data

        # ----------------------------------------------------
        # Create Adjustment Header
        # ----------------------------------------------------
        adjustment = StockAdjustment.objects.create(

            reason=data["reason"],

            note=data.get("note", ""),

            created_by=request.user
            if request.user.is_authenticated
            else None,

        )

        # ----------------------------------------------------
        # Create Adjustment Items
        # ----------------------------------------------------
        for item in data["items"]:

            product = Product.objects.get(
                id=item["product_id"]
            )

            quantity = item["quantity"]

            adjustment_type = item["adjustment_type"]

            previous_stock = product.stock_quantity

            if adjustment_type == "increase":

                stock_change = quantity

            else:

                if previous_stock < quantity:

                    raise serializers.ValidationError(

                        f"Insufficient stock for {product.name}"

                    )

                stock_change = -quantity

            # ---------------------------------------
            # Update Inventory
            # ---------------------------------------
            update_stock(

                product=product,

                quantity=stock_change,

                movement_type="adjustment",

                reference=adjustment.adjustment_number,

                note=item.get("note", ""),

            )

            product.refresh_from_db()

            # ---------------------------------------
            # Save Item
            # ---------------------------------------
            StockAdjustmentItem.objects.create(

                adjustment=adjustment,

                product=product,

                quantity=quantity,

                previous_stock=previous_stock,

                new_stock=product.stock_quantity,

                note=item.get("note", ""),

            )

        return Response({

            "success": True,

            "adjustment_number": adjustment.adjustment_number,

            "message": "Stock adjustment created successfully."

        })


class StockAdjustmentCreateView(APIView):

    @transaction.atomic
    def post(self, request):

        serializer = StockAdjustmentSerializer(
            data=request.data
        )

        serializer.is_valid(
            raise_exception=True
        )

        data = serializer.validated_data

        adjustment = StockAdjustment.objects.create(
            reason=data["reason"],
            note=data.get("note", ""),
            created_by=request.user
            if request.user.is_authenticated
            else None,
        )

        for item in data["items"]:

            product = get_object_or_404(
                Product,
                id=item["product_id"]
            )

            qty = item["quantity"]

            movement = item["adjustment_type"]
            
            if (
                movement == "decrease"
                and product.stock_quantity < qty
            ):
                raise serializers.ValidationError(
                    f"Not enough stock for {product.name}"
                )

        update_stock(
            product=product,
            quantity=qty if movement == "increase" else -qty,
            movement_type="adjustment",
            reference=adjustment.adjustment_number,
            note=item.get("note", ""),
        )                    

        product.refresh_from_db()

        StockAdjustmentItem.objects.create(
        
            adjustment=adjustment,

            product=product,

            adjustment_type=movement,

            quantity=qty,

            previous_stock=(
                product.stock_quantity - qty
                if movement == "increase"
                else product.stock_quantity + qty
            ),

            new_stock=product.stock_quantity,

            note=item.get("note", ""),

        )

        return Response({
        
            "success": True,

            "adjustment_number":
                adjustment.adjustment_number,

            "message":
                "Stock adjustment completed."

        })
    

class StockAdjustmentListView(APIView):

    def get(self, request):

        queryset = StockAdjustment.objects.prefetch_related(
            "items",
            "created_by",
        ).order_by("-created_at")

        # --------------------------------------------
        # Search
        # --------------------------------------------
        search = request.GET.get("search")

        if search:
            queryset = queryset.filter(
                Q(adjustment_number__icontains=search) |
                Q(reason__icontains=search) |
                Q(note__icontains=search)
            )

        # --------------------------------------------
        # Date Filter
        # --------------------------------------------
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        if start_date:
            queryset = queryset.filter(
                created_at__date__gte=start_date
            )

        if end_date:
            queryset = queryset.filter(
                created_at__date__lte=end_date
            )

        ordering = request.GET.get(
            "ordering",
            "-created_at"
        )

        queryset = queryset.order_by(ordering)

        paginator = CustomPagination()

        page = paginator.paginate_queryset(
            queryset,
            request
        )

        serializer = StockAdjustmentListSerializer(
            page,
            many=True
        )

        return paginator.get_paginated_response(
            serializer.data
        )


class StockAdjustmentDetailView(APIView):

    def get(self, request, adjustment_number):

        adjustment = get_object_or_404(

            StockAdjustment.objects.prefetch_related(
                "items__product"
            ),

            adjustment_number=adjustment_number

        )

        serializer = StockAdjustmentDetailSerializer(
            adjustment
        )

        return Response(serializer.data)  


class StockMovementDetailView(APIView):

    def get(self, request, pk):

        movement = get_object_or_404(
            StockMovement,
            pk=pk
        )

        serializer = StockMovementSerializer(movement)

        return Response(serializer.data)         
    

class StockAdjustmentExportCSVView(APIView):

    def get(self, request):

        adjustments = get_filtered_stock_adjustments(request)

        response = HttpResponse(
            content_type="text/csv"
        )

        filename = build_export_filename(
            request,
            "stock-adjustments",
            "csv"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        writer = csv.writer(response)

        writer.writerow([
            "S.No",
            "Adjustment No",
            "Reason",
            "Items",
            "Created By",
            "Date",
        ])

        for index, adjustment in enumerate(
            adjustments,
            start=1
        ):

            writer.writerow([

                index,

                adjustment.adjustment_number,

                adjustment.reason,

                adjustment.items.count(),

                adjustment.created_by.username
                if adjustment.created_by
                else "",

                timezone.localtime(
                    adjustment.created_at
                ).strftime("%d-%b-%Y %I:%M %p"),

            ])

        return response 


class StockAdjustmentExportExcelView(APIView):

    def get(self, request):

        adjustments = get_filtered_stock_adjustments(request)

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Stock Adjustments"

        headers = [
        
            "S.No",

            "Adjustment No",

            "Reason",

            "Items",

            "Created By",

            "Date",

        ]

        sheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="2E7D32"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        thin_border = Border(
        
            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin"),
        )

        for cell in sheet[1]:
        
            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = header_alignment

            cell.border = thin_border
            

        for index, adjustment in enumerate(adjustments, start=1):

            sheet.append([
            
                index,

                adjustment.adjustment_number,

                adjustment.reason,

                adjustment.items.count(),

                adjustment.created_by.username
                if adjustment.created_by
                else "",

                adjustment.created_at.strftime("%d-%b-%Y %I:%M %p"),

            ])

        for row in sheet.iter_rows(
            min_row=2
        ):

            for cell in row:
            
                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="center"
                )


        for column in sheet.columns:
        
            length = max(
                len(str(cell.value or ""))
                for cell in column
            )

            sheet.column_dimensions[
                column[0].column_letter
            ].width = length + 3

        sheet.freeze_panes = "A2"

        response = HttpResponse(
        
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        filename = build_export_filename(
            request,
            "stock-adjustments",
            "xlsx"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        workbook.save(response)

        return response


class StockAdjustmentExportPDFView(APIView):

    def get(self, request):

        buffer = BytesIO()

        doc = SimpleDocTemplate(buffer)

        styles = getSampleStyleSheet()

        story = []

        company = Company.objects.first()

        adjustments = get_filtered_stock_adjustments(request)    

        add_company_header(
            story,
            company,
        )
        
        add_report_title(
            story,
            "STOCK ADJUSTMENT REPORT"
        )
        
        story.append(
            Spacer(1, 15)
        )
        
        report_info = [
        
            ["Generated", timezone.localtime().strftime("%d-%b-%Y %I:%M %p")]
        
        ]

        search = request.GET.get("search")
        date = request.GET.get("date")

        if search:
            report_info.append(["Search", search])

        if date:
            report_info.append(["Date Filter", date.title()])

        add_report_info(
            story,
            report_info
        )            

        table_data = [[
        
            "S.No",

            "Adjustment Number",

            "Reason",

            "Items",

            "Created by",

            "Created at",

        ]]

        for index, adjustment in enumerate(adjustments, start=1):
        

            table_data.append([
            
                index,

                adjustment.adjustment_number,

                adjustment.reason,

                adjustment.items.count(),

                adjustment.created_by.username
                if adjustment.created_by
                else "",

                adjustment.created_at.strftime("%d-%b-%Y %I:%M %p"),

            ])

        table = Table(
        
            table_data,

            colWidths=[35,100,160,55,80,120],

            repeatRows=1

        )

        style_report_table(
            table,
            len(table_data),
            "#2E7D32"
        )

        story.append(table)

        summary = get_adjustment_summary(
            adjustments
        )

        summary_data = [
        
            [
                "Total Adjustments",
                summary["total_adjustments"]
            ],

            [
                "Products Adjusted",
                summary["total_products"]
            ],

            [
                "Increase Operations",
                summary["increase_operations"]
            ],

            [
                "Decrease Operations",
                summary["decrease_operations"]
            ],

            [
                "Total Quantity Adjusted",
                summary["total_quantity_adjusted"]
            ],

        ]

        add_summary_table(
            story,
            "SUMMARY",
            summary_data
        )

        doc.build(

            story,

            onFirstPage=add_page_footer,

            onLaterPages=add_page_footer,

        )

        pdf = buffer.getvalue()
        
        buffer.close()
        
        response = HttpResponse(
            pdf,
            content_type="application/pdf"
        )
        
        filename = build_export_filename(
            request,
            "stock-adjustments",
            "pdf"
        )
        
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        
        return response    


class InventoryValuationView(APIView):

    def get(self, request):

        data = get_inventory_valuation()

        serializer = InventoryValuationSerializer(data)

        return Response(serializer.data)   


class ProductInventoryValuationView(APIView):

    def get(self, request):

        products = get_filtered_inventory_valuation(request)

        serializer = ProductInventoryValuationSerializer(
            products,
            many=True,
        )

        return Response(serializer.data)


class ProductInventoryValuationView(APIView):

    def get(self, request):

        products = get_filtered_inventory_valuation(request)

        paginator = CustomPagination()

        page = paginator.paginate_queryset(
            products,
            request
        )

        serializer = ProductInventoryValuationSerializer(
            page,
            many=True
        )

        return paginator.get_paginated_response(
            serializer.data
        ) 


class InventoryReportView(APIView):

    def get(self, request):

        report = get_inventory_report(request)

        summary = InventoryReportSerializer(
            report["summary"]
        ).data

        products = ProductInventoryValuationSerializer(
            report["products"],
            many=True
        ).data

        return Response({

            "summary": summary,

            "products": products,

        }) 


class InventoryAnalyticsView(APIView):

    def get(self, request):

        analytics = get_inventory_analytics()

        serializer = InventoryAnalyticsSerializer(
            analytics
        )

        return Response(serializer.data)


class TopSellingProductsView(APIView):

    def get(self, request):

        limit = int(
            request.GET.get(
                "limit",
                10
            )
        )
        date_filter = request.GET.get(
            "date"
        )

        products = get_top_selling_products(
            limit,
            date_filter
        )

        serializer = (
            TopSellingProductSerializer(
                products,
                many=True
            )
        )

        return Response(
            serializer.data
        )


class SlowMovingProductsView(APIView):

    def get(self, request):

        limit = int(
            request.GET.get("limit", 10)
        )

        date_filter = request.GET.get("date")

        products = get_slow_moving_products(
            limit,
            date_filter,
        )

        serializer = SlowMovingProductSerializer(
            products,
            many=True
        )

        return Response(serializer.data)


class MostReturnedProductsView(APIView):

    def get(self, request):

        limit = int(
            request.GET.get("limit", 10)
        )

        date_filter = request.GET.get("date")

        products = get_most_returned_products(
            limit,
            date_filter,
        )

        serializer = MostReturnedProductSerializer(
            products,
            many=True,
        )

        return Response(serializer.data)  


class InventoryProductExportCSVView(APIView):

    def get(self, request):

        products = get_filtered_inventory(request)

        response = HttpResponse(
            content_type="text/csv"
        )

        filename = build_export_filename(
            request,
            "inventory",
            "csv"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        writer = csv.writer(response)

        writer.writerow([
            "S.No",
            "Product",
            "SKU",
            "Barcode",
            "Category",
            "Cost Price",
            "Sale Price",
            "Stock",
            "Min Stock",
            "Max Stock",
            "Unit",
            "Status",
            "Active",
        ])

        for index, product in enumerate(
            products,
            start=1
        ):

            if product.stock_quantity == 0:
                stock_status = "Out of Stock"

            elif product.stock_quantity <= product.min_stock:
                stock_status = "Low Stock"

            else:
                stock_status = "In Stock"

            writer.writerow([

                index,

                product.name,

                product.sku,

                product.barcode or "",

                product.category.name
                if product.category
                else "",

                product.cost_price,

                product.sales_price,

                product.stock_quantity,

                product.min_stock,

                product.max_stock or "",

                product.unit,

                stock_status,

                "Active"
                if product.is_active
                else "Inactive",

            ])

        return response
    

class InventoryProductExportExcelView(APIView):

    def get(self, request):

        products = get_filtered_inventory(request)

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Inventory"

        headers = [

            "S.No",

            "Product",

            "SKU",

            "Barcode",

            "Category",

            "Cost Price",

            "Sale Price",

            "Stock",

            "Min Stock",

            "Max Stock",

            "Unit",

            "Stock Status",

            "Active",

        ]

        sheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="2E7D32"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        thin_border = Border(

            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin"),

        )

        for cell in sheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = header_alignment

            cell.border = thin_border

        for index, product in enumerate(products, start=1):

            if product.stock_quantity == 0:

                stock_status = "Out of Stock"

            elif product.stock_quantity <= product.min_stock:

                stock_status = "Low Stock"

            else:

                stock_status = "In Stock"

            sheet.append([

                index,

                product.name,

                product.sku,

                product.barcode or "",

                product.category.name
                if product.category
                else "",

                product.cost_price,

                product.sales_price,

                product.stock_quantity,

                product.min_stock,

                product.max_stock or "",

                product.unit,

                stock_status,

                "Active"
                if product.is_active
                else "Inactive",

            ])

        for row in sheet.iter_rows(
            min_row=2
        ):

            for cell in row:

                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="center"
                )

        for column in sheet.columns:

            length = max(
                len(str(cell.value or ""))
                for cell in column
            )

            sheet.column_dimensions[
                column[0].column_letter
            ].width = length + 3

        sheet.freeze_panes = "A2"

        response = HttpResponse(

            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )

        )

        filename = build_export_filename(
            request,
            "inventory",
            "xlsx"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        workbook.save(response)

        return response   


class InventoryProductExportPDFView(APIView):

    def get(self, request):

        buffer = BytesIO()

        doc = SimpleDocTemplate(buffer)

        story = []

        company = Company.objects.first()

        products = get_filtered_inventory(request)

        add_company_header(
            story,
            company,
        )

        add_report_title(
            story,
            "INVENTORY REPORT"
        )

        story.append(
            Spacer(1, 15)
        )

        report_info = [

            [
                "Generated",
                timezone.localtime().strftime("%d-%b-%Y %I:%M %p")
            ]

        ]

        search = request.GET.get("search")

        status = request.GET.get("status")

        category = request.GET.get("category")

        if search:
            report_info.append([
                "Search",
                search
            ])

        if status:
            report_info.append([
                "Status",
                status.title()
            ])

        if category:
            category_name = Category.objects.filter(
                id=category
            ).values_list(
                "name",
                flat=True
            ).first()

            if category_name:
                report_info.append([
                    "Category",
                    category_name
                ])

        add_report_info(
            story,
            report_info
        )

        table_data = [[

            "S.No",

            "Product",

            "Category",

            "Stock",

            "Cost",

            "Sale",

            "Status",

        ]]

        for index, product in enumerate(products, start=1):

            if product.stock_quantity == 0:

                stock_status = "Out"

            elif product.stock_quantity <= product.min_stock:

                stock_status = "Low"

            else:

                stock_status = "In"

            table_data.append([

                index,

                product.name,

                product.category.name
                if product.category
                else "-",

                product.stock_quantity,

                f"Rs {product.cost_price}",

                f"Rs {product.sales_price}",

                stock_status,

            ])

        table = Table(

            table_data,

            colWidths=[
                35,
                140,
                90,
                55,
                65,
                65,
                60,
            ],

            repeatRows=1

        )

        style_report_table(

            table,

            len(table_data),

            "#2E7D32"

        )

        story.append(table)

        total_products = products.count()

        total_stock = sum(
            p.stock_quantity
            for p in products
        )

        inventory_cost = sum(
            p.stock_quantity * p.cost_price
            for p in products
        )

        inventory_sale = sum(
            p.stock_quantity * p.sales_price
            for p in products
        )

        summary = [

            [
                "Total Products",
                str(total_products)
            ],

            [
                "Total Stock Units",
                str(total_stock)
            ],

            [
                "Inventory Cost",
                f"Rs {inventory_cost:.2f}"
            ],

            [
                "Inventory Sale Value",
                f"Rs {inventory_sale:.2f}"
            ],

        ]

        add_summary_table(

            story,

            "SUMMARY",

            summary

        )

        doc.build(

            story,

            onFirstPage=add_page_footer,

            onLaterPages=add_page_footer,

        )

        pdf = buffer.getvalue()

        buffer.close()

        response = HttpResponse(

            pdf,

            content_type="application/pdf"

        )

        filename = build_export_filename(

            request,

            "inventory",

            "pdf"

        )

        response["Content-Disposition"] = (

            f'attachment; filename="{filename}"'

        )

        return response     
    

class ProductUpdateView(APIView):

    def put(self, request, pk):

        product = get_object_or_404(Product, pk=pk)

        serializer = ProductSerializer(
            product,
            data=request.data,
            partial=True
        )

        serializer.is_valid(raise_exception=True)

        serializer.save()

        return Response({
            "success": True,
            "message": "Product updated successfully.",
            "product": serializer.data
        })    


