"""
Views for Sales, Product, Stock, and Tax Reports.
These are NEW views — existing views are untouched.
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Sum, Count, Avg, Q, F
from django.utils import timezone
from datetime import timedelta
from pos.models import Order, OrderItem, Product, Category, Refund, Company
from ..pagination import CustomPagination
from ..pdf_helpers import (
    add_company_header, add_report_title, add_report_info,
    add_summary_table, style_report_table, 
)
from ..helpers import build_export_filename
from django.http import HttpResponse
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, Spacer, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


# ─── Shared date-filter helper ────────────────────────────────────────────────

def apply_date_filters(qs, request, date_field="created_at"):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    date = request.GET.get("date")
    today = timezone.now().date()

    if start_date:
        qs = qs.filter(**{f"{date_field}__date__gte": start_date})
    if end_date:
        qs = qs.filter(**{f"{date_field}__date__lte": end_date})

    if date == "today":
        qs = qs.filter(**{f"{date_field}__date": today})
    elif date == "yesterday":
        qs = qs.filter(**{f"{date_field}__date": today - timedelta(days=1)})
    elif date == "week":
        start_week = today - timedelta(days=today.weekday())
        qs = qs.filter(**{f"{date_field}__date__gte": start_week})
    elif date == "month":
        qs = qs.filter(**{f"{date_field}__year": today.year, f"{date_field}__month": today.month})
    elif date == "year":
        qs = qs.filter(**{f"{date_field}__year": today.year})

    return qs


# ─── Sales Report ─────────────────────────────────────────────────────────────

class SalesReportView(APIView):

    def get(self, request):
        orders = Order.objects.filter(status__in=["paid", "partially_refunded", "refunded"])
        orders = apply_date_filters(orders, request)

        payment = request.GET.get("payment")
        if payment:
            orders = orders.filter(payment_method=payment)

        agg = orders.aggregate(
            total_orders=Count("id"),
            total_revenue=Sum("total"),
            total_subtotal=Sum("subtotal"),
            total_discount=Sum("discount"),
            total_tax=Sum("tax"),
        )

        total_orders = agg["total_orders"] or 0
        total_revenue = float(agg["total_revenue"] or 0)
        total_discount = float(agg["total_discount"] or 0)
        total_tax = float(agg["total_tax"] or 0)
        avg_order_value = round(total_revenue / total_orders, 2) if total_orders > 0 else 0

        refunds_agg = Refund.objects.filter(
            order__in=orders
        ).aggregate(total_refunds=Sum("total_amount"))
        total_refunds = float(refunds_agg["total_refunds"] or 0)

        # Daily chart data (last 30 days by default)
        from django.db.models.functions import TruncDate
        chart_data = (
            orders
            .annotate(date=TruncDate("created_at"))
            .values("date")
            .annotate(revenue=Sum("total"), order_count=Count("id"))
            .order_by("date")
        )
        chart_list = [
            {
                "date": str(row["date"]),
                "revenue": float(row["revenue"] or 0),
                "order_count": row["order_count"],
            }
            for row in chart_data
        ]

        # Payment breakdown
        payment_breakdown = (
            orders
            .values("payment_method")
            .annotate(count=Count("id"), total=Sum("total"))
            .order_by("-total")
        )

        # Orders table (paginated)
        paginator = CustomPagination()
        page = paginator.paginate_queryset(orders.order_by("-created_at"), request)
        orders_data = [
            {
                "id": o.id,
                "order_number": o.order_number,
                "customer": o.customer.name if o.customer else "Walk-in Customer",
                "payment_method": o.payment_method,
                "subtotal": float(o.subtotal),
                "discount": float(o.discount),
                "tax": float(o.tax),
                "total": float(o.total),
                "status": o.status,
                "created_at": o.created_at,
            }
            for o in (page or [])
        ]

        response_data = {
            "summary": {
                "total_orders": total_orders,
                "total_revenue": total_revenue,
                "total_discount": total_discount,
                "total_tax": total_tax,
                "avg_order_value": avg_order_value,
                "total_refunds": total_refunds,
            },
            "chart_data": chart_list,
            "payment_breakdown": list(payment_breakdown),
        }

        if page is not None:
            paged = paginator.get_paginated_response(orders_data)
            paged.data["summary"] = response_data["summary"]
            paged.data["chart_data"] = chart_list
            paged.data["payment_breakdown"] = list(payment_breakdown)
            return paged

        response_data["results"] = orders_data
        return Response(response_data)


class SalesReportExportCSVView(APIView):

    def get(self, request):
        orders = Order.objects.filter(status__in=["paid", "partially_refunded", "refunded"])
        orders = apply_date_filters(orders, request)
        payment = request.GET.get("payment")
        if payment:
            orders = orders.filter(payment_method=payment)
        orders = orders.order_by("-created_at")

        response = HttpResponse(content_type="text/csv")
        filename = build_export_filename(request, "sales_report", "csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(["Order No", "Customer", "Payment", "Subtotal", "Discount", "Tax", "Total", "Status", "Date"])
        for o in orders:
            writer.writerow([
                o.order_number,
                o.customer.name if o.customer else "Walk-in Customer",
                o.payment_method,
                o.subtotal, o.discount, o.tax, o.total, o.status,
                o.created_at.strftime("%Y-%m-%d %H:%M"),
            ])
        return response


class SalesReportExportExcelView(APIView):

    def get(self, request):
        orders = Order.objects.filter(status__in=["paid", "partially_refunded", "refunded"])
        orders = apply_date_filters(orders, request).order_by("-created_at")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        headers = ["Order No", "Customer", "Payment", "Subtotal", "Discount", "Tax", "Total", "Status", "Date"]
        ws.append(headers)
        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for o in orders:
            ws.append([
                o.order_number,
                o.customer.name if o.customer else "Walk-in Customer",
                o.payment_method,
                float(o.subtotal), float(o.discount), float(o.tax), float(o.total),
                o.status, o.created_at.strftime("%Y-%m-%d %H:%M"),
            ])
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value)) if c.value else 0 for c in col
            ) + 3
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = build_export_filename(request, "sales_report", "xlsx")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class SalesReportExportPDFView(APIView):

    def get(self, request):
        orders = Order.objects.filter(status__in=["paid", "partially_refunded", "refunded"])
        orders = apply_date_filters(orders, request).order_by("-created_at")
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        company = Company.objects.first()
        story = []
        add_company_header(story, company)
        add_report_title(story, "SALES REPORT")
        generated = timezone.localtime().strftime("%d-%b-%Y %I:%M %p")
        add_report_info(story, [["Generated", generated]])

        total_revenue = 0
        total_tax = 0
        total_discount = 0
        table_data = [["#", "Order No", "Customer", "Payment", "Total", "Status", "Date"]]
        for i, o in enumerate(orders, 1):
            total_revenue += float(o.total)
            total_tax += float(o.tax)
            total_discount += float(o.discount)
            table_data.append([
                i, o.order_number,
                o.customer.name if o.customer else "Walk-in Customer",
                o.payment_method,
                f"Rs {o.total:,.2f}",
                o.status.replace("_", " ").title(),
                o.created_at.strftime("%d-%b-%Y"),
            ])
        table = Table(table_data)
        style_report_table(table, len(table_data))
        story.append(table)
        story.append(Spacer(1, 20))
        add_summary_table(story, "SUMMARY", [
            ["Total Orders", str(orders.count())],
            ["Total Revenue", f"Rs {total_revenue:,.2f}"],
            ["Total Discount", f"Rs {total_discount:,.2f}"],
            ["Total Tax", f"Rs {total_tax:,.2f}"],
        ])
        doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        filename = build_export_filename(request, "sales_report", "pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# ─── Product Report ───────────────────────────────────────────────────────────

class ProductReportView(APIView):

    def get(self, request):
        items = OrderItem.objects.select_related("product", "product__category")
        items = apply_date_filters(items, request, "order__created_at")
        items = items.filter(order__status__in=["paid", "partially_refunded", "refunded"])

        category = request.GET.get("category")
        if category:
            items = items.filter(product__category_id=category)

        product_stats = (
            items
            .values("product__id", "product__name", "product__sku", "product__category__name")
            .annotate(
                qty_sold=Sum("quantity"),
                revenue=Sum("subtotal"),
            )
            .order_by("-revenue")
        )

        products_data = []
        for row in product_stats:
            # Refund count
            from ..models import RefundItem
            refund_count = RefundItem.objects.filter(
                order_item__product_id=row["product__id"]
            ).aggregate(total=Sum("quantity"))["total"] or 0
            try:
                p = Product.objects.get(id=row["product__id"])
                cost = float(p.cost_price) * float(row["qty_sold"] or 0)
                profit = float(row["revenue"] or 0) - cost
            except Product.DoesNotExist:
                profit = 0
            products_data.append({
                "product_id": row["product__id"],
                "product_name": row["product__name"],
                "sku": row["product__sku"],
                "category": row["product__category__name"] or "-",
                "qty_sold": float(row["qty_sold"] or 0),
                "revenue": float(row["revenue"] or 0),
                "return_count": refund_count,
                "profit": round(profit, 2),
            })

        agg = items.aggregate(
            total_qty=Sum("quantity"),
            total_revenue=Sum("subtotal"),
        )
        summary = {
            "total_products_sold": len(products_data),
            "total_qty_sold": float(agg["total_qty"] or 0),
            "total_revenue": float(agg["total_revenue"] or 0),
            "top_product": products_data[0]["product_name"] if products_data else "-",
        }

        paginator = CustomPagination()
        page = paginator.paginate_queryset(products_data, request)
        paged = paginator.get_paginated_response(page or products_data)
        paged.data["summary"] = summary
        return paged


class ProductReportExportCSVView(APIView):

    def get(self, request):
        items = OrderItem.objects.select_related("product", "product__category")
        items = apply_date_filters(items, request, "order__created_at")
        items = items.filter(order__status__in=["paid", "partially_refunded", "refunded"])
        product_stats = (
            items.values("product__name", "product__sku", "product__category__name")
            .annotate(qty_sold=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-revenue")
        )
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="product_report.csv"'
        writer = csv.writer(response)
        writer.writerow(["Product", "SKU", "Category", "Qty Sold", "Revenue"])
        for row in product_stats:
            writer.writerow([
                row["product__name"], row["product__sku"],
                row["product__category__name"] or "-",
                row["qty_sold"], row["revenue"],
            ])
        return response


class ProductReportExportExcelView(APIView):

    def get(self, request):
        items = OrderItem.objects.select_related("product", "product__category")
        items = apply_date_filters(items, request, "order__created_at")
        items = items.filter(order__status__in=["paid", "partially_refunded", "refunded"])
        product_stats = (
            items.values("product__name", "product__sku", "product__category__name")
            .annotate(qty_sold=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-revenue")
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Report"
        ws.append(["Product", "SKU", "Category", "Qty Sold", "Revenue"])
        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
        for row in product_stats:
            ws.append([
                row["product__name"], row["product__sku"],
                row["product__category__name"] or "-",
                float(row["qty_sold"] or 0), float(row["revenue"] or 0),
            ])
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value)) if c.value else 0 for c in col
            ) + 3
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="product_report.xlsx"'
        wb.save(response)
        return response


class ProductReportExportPDFView(APIView):

    def get(self, request):
        items = OrderItem.objects.select_related("product")
        items = apply_date_filters(items, request, "order__created_at")
        items = items.filter(order__status__in=["paid", "partially_refunded", "refunded"])
        product_stats = (
            items.values("product__name", "product__sku", "product__category__name")
            .annotate(qty_sold=Sum("quantity"), revenue=Sum("subtotal"))
            .order_by("-revenue")
        )
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        company = Company.objects.first()
        story = []
        add_company_header(story, company)
        add_report_title(story, "PRODUCT PERFORMANCE REPORT")
        add_report_info(story, [["Generated", timezone.localtime().strftime("%d-%b-%Y %I:%M %p")]])
        table_data = [["#", "Product", "SKU", "Category", "Qty Sold", "Revenue"]]
        for i, row in enumerate(product_stats, 1):
            table_data.append([
                i, row["product__name"], row["product__sku"],
                row["product__category__name"] or "-",
                row["qty_sold"], f"Rs {row['revenue']:,.2f}",
            ])
        table = Table(table_data)
        style_report_table(table, len(table_data))
        story.append(table)
        doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="product_report.pdf"'
        return response


# ─── Stock Report ─────────────────────────────────────────────────────────────

class StockReportView(APIView):

    def get(self, request):
        products = Product.objects.select_related("category", "brand")

        search = request.GET.get("search", "").strip()
        if search:
            products = products.filter(Q(name__icontains=search) | Q(sku__icontains=search))

        category = request.GET.get("category")
        if category:
            products = products.filter(category_id=category)

        stock_status = request.GET.get("status")
        if stock_status == "low":
            products = products.filter(stock_quantity__gt=0, stock_quantity__lte=F("min_stock"))
        elif stock_status == "out":
            products = products.filter(stock_quantity=0)
        elif stock_status == "in":
            products = products.filter(stock_quantity__gt=F("min_stock"))
        elif stock_status == "active":
            products = products.filter(is_active=True)

        products = products.order_by("name")

        summary_qs = products
        total_products = summary_qs.count()
        low_stock = summary_qs.filter(stock_quantity__gt=0, stock_quantity__lte=F("min_stock")).count()
        out_of_stock = summary_qs.filter(stock_quantity=0).count()
        total_value = sum(
            float(p.cost_price) * p.stock_quantity for p in summary_qs
        )

        paginator = CustomPagination()
        page = paginator.paginate_queryset(products, request)

        products_data = [
            {
                "id": p.id,
                "name": p.name,
                "sku": p.sku,
                "category": p.category.name if p.category else "-",
                "brand": p.brand.name if p.brand else "-",
                "stock_quantity": p.stock_quantity,
                "min_stock": p.min_stock,
                "max_stock": p.max_stock,
                "cost_price": float(p.cost_price),
                "sales_price": float(p.sales_price),
                "inventory_value": float(p.cost_price) * p.stock_quantity,
                "stock_status": (
                    "Out of Stock" if p.stock_quantity == 0 else
                    "Low Stock" if p.stock_quantity <= p.min_stock else
                    "In Stock"
                ),
                "is_active": p.is_active,
            }
            for p in (page or [])
        ]

        paged = paginator.get_paginated_response(products_data)
        paged.data["summary"] = {
            "total_products": total_products,
            "low_stock": low_stock,
            "out_of_stock": out_of_stock,
            "total_value": round(total_value, 2),
        }
        return paged


class StockReportExportCSVView(APIView):

    def get(self, request):
        products = Product.objects.select_related("category", "brand").order_by("name")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="stock_report.csv"'
        writer = csv.writer(response)
        writer.writerow(["Name", "SKU", "Category", "Brand", "Stock", "Min", "Max",
                         "Cost Price", "Sales Price", "Inventory Value", "Status"])
        for p in products:
            status = "Out of Stock" if p.stock_quantity == 0 else (
                "Low Stock" if p.stock_quantity <= p.min_stock else "In Stock"
            )
            writer.writerow([
                p.name, p.sku,
                p.category.name if p.category else "-",
                p.brand.name if p.brand else "-",
                p.stock_quantity, p.min_stock, p.max_stock or "-",
                p.cost_price, p.sales_price,
                float(p.cost_price) * p.stock_quantity, status,
            ])
        return response


class StockReportExportExcelView(APIView):

    def get(self, request):
        products = Product.objects.select_related("category", "brand").order_by("name")
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Report"
        headers = ["Name", "SKU", "Category", "Brand", "Stock", "Min Stock",
                   "Cost Price", "Sales Price", "Inventory Value", "Status"]
        ws.append(headers)
        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
        for p in products:
            status = "Out of Stock" if p.stock_quantity == 0 else (
                "Low Stock" if p.stock_quantity <= p.min_stock else "In Stock"
            )
            ws.append([
                p.name, p.sku,
                p.category.name if p.category else "-",
                p.brand.name if p.brand else "-",
                p.stock_quantity, p.min_stock,
                float(p.cost_price), float(p.sales_price),
                float(p.cost_price) * p.stock_quantity, status,
            ])
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value)) if c.value else 0 for c in col
            ) + 3
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="stock_report.xlsx"'
        wb.save(response)
        return response


class StockReportExportPDFView(APIView):

    def get(self, request):
        products = Product.objects.select_related("category").order_by("name")
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        company = Company.objects.first()
        story = []
        add_company_header(story, company)
        add_report_title(story, "STOCK REPORT")
        add_report_info(story, [["Generated", timezone.localtime().strftime("%d-%b-%Y %I:%M %p")]])
        table_data = [["#", "Product", "SKU", "Category", "Stock", "Min", "Cost Value", "Status"]]
        total_value = 0
        for i, p in enumerate(products, 1):
            val = float(p.cost_price) * p.stock_quantity
            total_value += val
            status = "Out" if p.stock_quantity == 0 else ("Low" if p.stock_quantity <= p.min_stock else "OK")
            table_data.append([
                i, p.name, p.sku,
                p.category.name if p.category else "-",
                p.stock_quantity, p.min_stock,
                f"Rs {val:,.2f}", status,
            ])
        table = Table(table_data)
        style_report_table(table, len(table_data))
        story.append(table)
        story.append(Spacer(1, 20))
        add_summary_table(story, "SUMMARY", [
            ["Total Products", str(products.count())],
            ["Total Inventory Value", f"Rs {total_value:,.2f}"],
        ])
        doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="stock_report.pdf"'
        return response


# ─── Tax Report ───────────────────────────────────────────────────────────────

class TaxReportView(APIView):

    def get(self, request):
        orders = Order.objects.filter(
            status__in=["paid", "partially_refunded", "refunded"],
            tax__gt=0,
        ).select_related("customer")
        orders = apply_date_filters(orders, request)

        payment = request.GET.get("payment")
        if payment:
            orders = orders.filter(payment_method=payment)

        orders = orders.order_by("-created_at")
        agg = orders.aggregate(
            total_tax=Sum("tax"),
            total_orders=Count("id"),
            avg_tax=Avg("tax"),
        )
        total_tax = float(agg["total_tax"] or 0)
        avg_tax = float(agg["avg_tax"] or 0)
        today = timezone.now()

        this_month_tax = float(
            orders.filter(created_at__year=today.year, created_at__month=today.month)
            .aggregate(t=Sum("tax"))["t"] or 0
        )
        start_week = today.date() - timedelta(days=today.weekday())
        this_week_tax = float(
            orders.filter(created_at__date__gte=start_week)
            .aggregate(t=Sum("tax"))["t"] or 0
        )

        paginator = CustomPagination()
        page = paginator.paginate_queryset(orders, request)
        orders_data = [
            {
                "order_number": o.order_number,
                "created_at": o.created_at,
                "customer": o.customer.name if o.customer else "Walk-in Customer",
                "subtotal": float(o.subtotal),
                "tax_rate": "18%",
                "tax": float(o.tax),
                "total": float(o.total),
                "payment_method": o.payment_method,
            }
            for o in (page or [])
        ]

        paged = paginator.get_paginated_response(orders_data)
        paged.data["summary"] = {
            "total_tax": total_tax,
            "this_month_tax": this_month_tax,
            "this_week_tax": this_week_tax,
            "avg_tax": round(avg_tax, 2),
            "total_orders": agg["total_orders"] or 0,
        }
        return paged


class TaxReportExportCSVView(APIView):

    def get(self, request):
        orders = Order.objects.filter(
            status__in=["paid", "partially_refunded", "refunded"], tax__gt=0
        ).select_related("customer")
        orders = apply_date_filters(orders, request).order_by("-created_at")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="tax_report.csv"'
        writer = csv.writer(response)
        writer.writerow(["Order No", "Date", "Customer", "Subtotal", "Tax", "Total", "Payment"])
        for o in orders:
            writer.writerow([
                o.order_number,
                o.created_at.strftime("%Y-%m-%d %H:%M"),
                o.customer.name if o.customer else "Walk-in Customer",
                o.subtotal, o.tax, o.total, o.payment_method,
            ])
        return response


class TaxReportExportExcelView(APIView):

    def get(self, request):
        orders = Order.objects.filter(
            status__in=["paid", "partially_refunded", "refunded"], tax__gt=0
        ).select_related("customer")
        orders = apply_date_filters(orders, request).order_by("-created_at")
        wb = Workbook()
        ws = wb.active
        ws.title = "Tax Report"
        ws.append(["Order No", "Date", "Customer", "Subtotal", "Tax", "Total", "Payment"])
        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
        for o in orders:
            ws.append([
                o.order_number,
                o.created_at.strftime("%Y-%m-%d %H:%M"),
                o.customer.name if o.customer else "Walk-in Customer",
                float(o.subtotal), float(o.tax), float(o.total), o.payment_method,
            ])
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value)) if c.value else 0 for c in col
            ) + 3
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="tax_report.xlsx"'
        wb.save(response)
        return response


class TaxReportExportPDFView(APIView):

    def get(self, request):
        orders = Order.objects.filter(
            status__in=["paid", "partially_refunded", "refunded"], tax__gt=0
        ).select_related("customer")
        orders = apply_date_filters(orders, request).order_by("-created_at")
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        company = Company.objects.first()
        story = []
        add_company_header(story, company)
        add_report_title(story, "TAX REPORT")
        add_report_info(story, [["Generated", timezone.localtime().strftime("%d-%b-%Y %I:%M %p")]])
        total_tax = 0
        table_data = [["#", "Order No", "Customer", "Subtotal", "Tax", "Total", "Date"]]
        for i, o in enumerate(orders, 1):
            total_tax += float(o.tax)
            table_data.append([
                i, o.order_number,
                o.customer.name if o.customer else "Walk-in Customer",
                f"Rs {o.subtotal:,.2f}", f"Rs {o.tax:,.2f}", f"Rs {o.total:,.2f}",
                o.created_at.strftime("%d-%b-%Y"),
            ])
        table = Table(table_data)
        style_report_table(table, len(table_data))
        story.append(table)
        story.append(Spacer(1, 20))
        add_summary_table(story, "SUMMARY", [
            ["Total Orders", str(orders.count())],
            ["Total Tax Collected", f"Rs {total_tax:,.2f}"],
        ])
        doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="tax_report.pdf"'
        return response
