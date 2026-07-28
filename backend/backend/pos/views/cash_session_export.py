import csv
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib.styles import getSampleStyleSheet

from ..models import CashSession, Company
from ..serializers import CashSessionSerializer
from ..helpers import build_export_filename
from .orders import (
    add_company_header, add_report_title, add_report_info, 
    style_report_table, add_summary_table, add_page_footer
)

def get_filtered_sessions(request):
    sessions = CashSession.objects.all().order_by("-id")
    status = request.GET.get("status")
    session_id = request.GET.get("id")
    date_filter = request.GET.get("date")
    cashier = request.GET.get("cashier")
    
    if status == "open":
        sessions = sessions.filter(is_open=True)
    elif status == "closed":
        sessions = sessions.filter(is_open=False)
        
    if session_id:
        sessions = sessions.filter(id=session_id)
        
    if date_filter and date_filter != 'all' and date_filter != 'today':
        sessions = sessions.filter(opened_at__date=date_filter)
    elif date_filter == 'today':
        from django.utils.timezone import now
        sessions = sessions.filter(opened_at__date=now().date())
        
    if cashier:
        sessions = sessions.filter(employee_name__icontains=cashier)
        
    search = request.GET.get("search")
    if search:
        sessions = sessions.filter(id__icontains=search)
        
        
    # More filters can be added here if passed (cashier, etc.)
    return sessions

class CashSessionExportCSVView(APIView):
    def get(self, request):
        sessions = get_filtered_sessions(request)
        response = HttpResponse(content_type="text/csv")
        filename = build_export_filename(request, "cash_sessions", "csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            "Session ID",
            "Date",
            "Cashier",
            "Opened At",
            "Closed At",
            "Opening Cash",
            "Expected Cash",
            "Actual Cash",
            "Difference",
            "Total Sales",
            "Status"
        ])

        serializer_data = CashSessionSerializer(sessions, many=True).data

        for s in serializer_data:
            writer.writerow([
                s.get("id"),
                s.get("opened_at")[:10],
                s.get("employee_name") or "Admin",
                s.get("opened_at"),
                s.get("closed_at") or "-",
                s.get("opening_balance"),
                s.get("expected_cash"),
                s.get("actual_closing_balance") or "-",
                s.get("difference"),
                s.get("total_sales"),
                "Open" if s.get("is_open") else "Closed"
            ])

        return response


class CashSessionExportExcelView(APIView):
    def get(self, request):
        sessions = get_filtered_sessions(request)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Cash Sessions"
        sheet.freeze_panes = "A2"

        headers = [
            "Session ID",
            "Date",
            "Cashier",
            "Opened At",
            "Closed At",
            "Opening Cash",
            "Expected Cash",
            "Actual Cash",
            "Difference",
            "Total Sales",
            "Status"
        ]
        sheet.append(headers)

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        serializer_data = CashSessionSerializer(sessions, many=True).data

        for s in serializer_data:
            sheet.append([
                s.get("id"),
                s.get("opened_at")[:10],
                s.get("employee_name") or "Admin",
                s.get("opened_at")[:19].replace('T', ' '),
                s.get("closed_at")[:19].replace('T', ' ') if s.get("closed_at") else "-",
                float(s.get("opening_balance") or 0),
                float(s.get("expected_cash") or 0),
                float(s.get("actual_closing_balance") or 0) if s.get("actual_closing_balance") is not None else 0.0,
                float(s.get("difference") or 0),
                float(s.get("total_sales") or 0),
                "Open" if s.get("is_open") else "Closed"
            ])

        sheet.auto_filter.ref = sheet.dimensions

        for col in ["F", "G", "H", "I", "J"]:
            for cell in sheet[col][1:]:
                cell.number_format = '#,##0.00'

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 3

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = build_export_filename(request, "cash_sessions", "xlsx")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response


class CashSessionExportPDFView(APIView):
    def get(self, request):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        styles = getSampleStyleSheet()

        company = Company.objects.first()
        story = []
        sessions = get_filtered_sessions(request)
        add_company_header(story, company)
        add_report_title(story, "CASH SESSIONS REPORT")

        generated = timezone.localtime().strftime("%d-%b-%Y %I:%M %p")
        story.append(Spacer(1, 10))

        status_filter = request.GET.get("status")
        report_info = [["Generated", generated]]
        if status_filter:
            report_info.append(["Status", status_filter.title()])
        add_report_info(story, report_info)

        if not sessions.exists():
            story.append(Paragraph("No cash sessions found for the selected filters.", styles["Heading3"]))

        table_data = [[
            "ID",
            "Cashier",
            "Expected",
            "Actual",
            "Diff",
            "Status",
            "Date",
        ]]

        serializer_data = CashSessionSerializer(sessions, many=True).data

        total_sessions = 0
        total_expected = 0
        total_actual = 0
        total_diff = 0

        for s in serializer_data:
            total_sessions += 1
            expected = float(s.get("expected_cash") or 0)
            actual = float(s.get("actual_closing_balance") or 0) if not s.get("is_open") else expected
            diff = float(s.get("difference") or 0)
            
            total_expected += expected
            total_actual += actual
            total_diff += diff

            table_data.append([
                str(s.get("id")),
                (s.get("employee_name") or "Admin")[:12],
                f"Rs {expected:,.2f}",
                f"Rs {actual:,.2f}",
                f"Rs {diff:,.2f}",
                "Open" if s.get("is_open") else "Closed",
                s.get("opened_at")[:10],
            ])

        table = Table(table_data)
        style_report_table(table, len(table_data))
        story.append(table)
        story.append(Spacer(1, 25))

        summary_data = [
            ["Total Sessions", str(total_sessions)],
            ["Total Expected Cash", f"Rs {total_expected:,.2f}"],
            ["Total Actual Cash", f"Rs {total_actual:,.2f}"],
            ["Total Difference", f"Rs {total_diff:,.2f}"],
        ]
        add_summary_table(story, "SUMMARY", summary_data)

        doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type="application/pdf")
        filename = build_export_filename(request, "cash_sessions", "pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
