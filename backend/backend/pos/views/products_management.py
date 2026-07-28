"""
Views for Products Management, Categories, Brands, and Variants.
These are NEW views — existing views are untouched.
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q, Count
from django.shortcuts import get_object_or_404
from pos.models import Product, Category, Brand, Variant, VariantValue
from ..pagination import CustomPagination
from ..helpers import build_export_filename
from ..pdf_helpers import (
    add_company_header, add_report_title, add_report_info,
    add_summary_table, style_report_table,
)
from pos.models import Company
import csv
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, Spacer
from reportlab.lib.styles import getSampleStyleSheet


# ─── Serializers inline (kept with views for self-containment) ────────────────
from rest_framework import serializers


class BrandSerializer(serializers.ModelSerializer):
    product_count = serializers.SerializerMethodField()

    class Meta:
        model = Brand
        fields = ["id", "name", "description", "logo", "is_active", "created_at", "product_count"]

    def get_product_count(self, obj):
        return obj.product_set.count()


class CategoryManageSerializer(serializers.ModelSerializer):
    product_count = serializers.SerializerMethodField()

    class Meta:
        model = Category
        fields = ["id", "name", "product_count"]

    def get_product_count(self, obj):
        return obj.product_set.count()


class ProductManageSerializer(serializers.ModelSerializer):
    category_name = serializers.CharField(source="category.name", read_only=True, default="")
    brand_name = serializers.CharField(source="brand.name", read_only=True, default="")
    stock_status = serializers.SerializerMethodField()
    image_url = serializers.SerializerMethodField()

    class Meta:
        model = Product
        fields = [
            "id", "name", "sku", "barcode", "description",
            "category", "category_name",
            "brand", "brand_name",
            "sales_price", "cost_price",
            "stock_quantity", "min_stock", "max_stock", "unit",
            "image", "image_url",
            "is_active", "stock_status",
        ]

    def get_stock_status(self, obj):
        if obj.stock_quantity <= 0:
            return "Out of Stock"
        if obj.stock_quantity <= obj.min_stock:
            return "Low Stock"
        return "In Stock"

    def get_image_url(self, obj):
        if obj.image:
            request = self.context.get("request")
            if request:
                return request.build_absolute_uri(obj.image.url)
            return obj.image.url
        return None


class VariantValueSerializer(serializers.ModelSerializer):
    class Meta:
        model = VariantValue
        fields = ["id", "value"]


class VariantSerializer(serializers.ModelSerializer):
    values = VariantValueSerializer(many=True, read_only=True)
    values_count = serializers.SerializerMethodField()

    class Meta:
        model = Variant
        fields = ["id", "name", "created_at", "values", "values_count"]

    def get_values_count(self, obj):
        return obj.values.count()


# ─── Products ─────────────────────────────────────────────────────────────────

class AllProductsView(APIView):

    def get(self, request):
        products = Product.objects.select_related("category", "brand").order_by("name")

        search = request.GET.get("search", "").strip()
        if search:
            products = products.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(barcode__icontains=search)
            )

        category = request.GET.get("category")
        if category:
            products = products.filter(category_id=category)

        brand = request.GET.get("brand")
        if brand:
            products = products.filter(brand_id=brand)

        status_filter = request.GET.get("status")
        if status_filter == "active":
            products = products.filter(is_active=True)
        elif status_filter == "inactive":
            products = products.filter(is_active=False)
        elif status_filter == "low":
            from django.db.models import F
            products = products.filter(stock_quantity__gt=0, stock_quantity__lte=F("min_stock"))
        elif status_filter == "out":
            products = products.filter(stock_quantity=0)

        paginator = CustomPagination()
        page = paginator.paginate_queryset(products, request)
        serializer = ProductManageSerializer(page, many=True, context={"request": request})
        return paginator.get_paginated_response(serializer.data)

    def post(self, request):
        serializer = ProductManageSerializer(data=request.data, context={"request": request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductManageDetailView(APIView):

    def get(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        serializer = ProductManageSerializer(product, context={"request": request})
        return Response(serializer.data)

    def put(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        serializer = ProductManageSerializer(
            product, data=request.data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        product.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ProductStatsView(APIView):
    """Quick stats for the All Products page cards."""

    def get(self, request):
        from django.db.models import F
        total = Product.objects.count()
        active = Product.objects.filter(is_active=True).count()
        low = Product.objects.filter(
            is_active=True,
            stock_quantity__gt=0,
            stock_quantity__lte=F("min_stock"),
        ).count()
        out = Product.objects.filter(is_active=True, stock_quantity=0).count()
        return Response({
            "total": total,
            "active": active,
            "low_stock": low,
            "out_of_stock": out,
        })


class AllProductsExportCSVView(APIView):

    def get(self, request):
        from django.db.models import F
        products = Product.objects.select_related("category", "brand").order_by("name")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="products.csv"'
        writer = csv.writer(response)
        writer.writerow(["Name", "SKU", "Barcode", "Category", "Brand",
                         "Sales Price", "Cost Price", "Stock", "Unit", "Status"])
        for p in products:
            writer.writerow([
                p.name, p.sku, p.barcode or "",
                p.category.name if p.category else "",
                p.brand.name if p.brand else "",
                p.sales_price, p.cost_price, p.stock_quantity, p.unit,
                "Active" if p.is_active else "Inactive",
            ])
        return response


class AllProductsExportExcelView(APIView):

    def get(self, request):
        products = Product.objects.select_related("category", "brand").order_by("name")
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["Name", "SKU", "Barcode", "Category", "Brand",
                   "Sales Price", "Cost Price", "Stock", "Unit", "Status"]
        ws.append(headers)
        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for p in products:
            ws.append([
                p.name, p.sku, p.barcode or "",
                p.category.name if p.category else "",
                p.brand.name if p.brand else "",
                float(p.sales_price), float(p.cost_price),
                p.stock_quantity, p.unit,
                "Active" if p.is_active else "Inactive",
            ])
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value)) if c.value else 0 for c in col
            ) + 3
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
        wb.save(response)
        return response


class AllProductsExportPDFView(APIView):

    def get(self, request):
        products = Product.objects.select_related("category", "brand").order_by("name")
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        styles = getSampleStyleSheet()
        company = Company.objects.first()
        story = []
        add_company_header(story, company)
        add_report_title(story, "ALL PRODUCTS REPORT")
        generated = timezone.localtime().strftime("%d-%b-%Y %I:%M %p")
        add_report_info(story, [["Generated", generated]])
        table_data = [["#", "Name", "SKU", "Category", "Brand", "Sales Price", "Stock", "Status"]]
        for i, p in enumerate(products, 1):
            table_data.append([
                i, p.name, p.sku,
                p.category.name if p.category else "-",
                p.brand.name if p.brand else "-",
                f"Rs {p.sales_price:,.2f}",
                p.stock_quantity,
                "Active" if p.is_active else "Inactive",
            ])
        table = Table(table_data)
        style_report_table(table, len(table_data))
        story.append(table)
        doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="products.pdf"'
        return response


# ─── Categories ───────────────────────────────────────────────────────────────

class CategoriesManageView(APIView):

    def get(self, request):
        categories = Category.objects.annotate(product_count=Count("product")).order_by("name")
        search = request.GET.get("search", "").strip()
        if search:
            categories = categories.filter(name__icontains=search)
        paginator = CustomPagination()
        page = paginator.paginate_queryset(categories, request)
        serializer = CategoryManageSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def post(self, request):
        serializer = CategoryManageSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CategoryManageDetailView(APIView):

    def get(self, request, pk):
        cat = get_object_or_404(Category, pk=pk)
        serializer = CategoryManageSerializer(cat)
        return Response(serializer.data)

    def put(self, request, pk):
        cat = get_object_or_404(Category, pk=pk)
        serializer = CategoryManageSerializer(cat, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        cat = get_object_or_404(Category, pk=pk)
        cat.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CategoryStatsView(APIView):
    def get(self, request):
        total = Category.objects.count()
        total_products = Product.objects.filter(is_active=True).count()
        return Response({"total": total, "total_products": total_products})


class CategoriesExportCSVView(APIView):
    def get(self, request):
        categories = Category.objects.annotate(product_count=Count("product")).order_by("name")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="categories.csv"'
        writer = csv.writer(response)
        writer.writerow(["Name", "Product Count"])
        for c in categories:
            writer.writerow([c.name, c.product_count])
        return response


# ─── Brands ───────────────────────────────────────────────────────────────────

class BrandsView(APIView):

    def get(self, request):
        brands = Brand.objects.annotate(product_count=Count("product")).order_by("name")
        search = request.GET.get("search", "").strip()
        if search:
            brands = brands.filter(Q(name__icontains=search) | Q(description__icontains=search))
        is_active = request.GET.get("is_active")
        if is_active == "true":
            brands = brands.filter(is_active=True)
        elif is_active == "false":
            brands = brands.filter(is_active=False)
        paginator = CustomPagination()
        page = paginator.paginate_queryset(brands, request)
        serializer = BrandSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def post(self, request):
        serializer = BrandSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BrandDetailView(APIView):

    def get(self, request, pk):
        brand = get_object_or_404(Brand, pk=pk)
        serializer = BrandSerializer(brand)
        return Response(serializer.data)

    def put(self, request, pk):
        brand = get_object_or_404(Brand, pk=pk)
        serializer = BrandSerializer(brand, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        brand = get_object_or_404(Brand, pk=pk)
        brand.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BrandStatsView(APIView):
    def get(self, request):
        total = Brand.objects.count()
        active = Brand.objects.filter(is_active=True).count()
        inactive = Brand.objects.filter(is_active=False).count()
        total_products = Product.objects.filter(brand__isnull=False).count()
        return Response({
            "total": total, "active": active,
            "inactive": inactive, "total_products": total_products,
        })


# ─── Variants ─────────────────────────────────────────────────────────────────

class VariantsView(APIView):

    def get(self, request):
        variants = Variant.objects.prefetch_related("values").order_by("name")
        search = request.GET.get("search", "").strip()
        if search:
            variants = variants.filter(name__icontains=search)
        paginator = CustomPagination()
        page = paginator.paginate_queryset(variants, request)
        serializer = VariantSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def post(self, request):
        name = request.data.get("name", "").strip()
        if not name:
            return Response({"error": "Name is required"}, status=status.HTTP_400_BAD_REQUEST)
        variant = Variant.objects.create(name=name)
        values = request.data.get("values", [])
        for val in values:
            v = val.strip() if isinstance(val, str) else val.get("value", "").strip()
            if v:
                VariantValue.objects.get_or_create(variant=variant, value=v)
        serializer = VariantSerializer(variant)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class VariantDetailView(APIView):

    def get(self, request, pk):
        variant = get_object_or_404(Variant, pk=pk)
        serializer = VariantSerializer(variant)
        return Response(serializer.data)

    def put(self, request, pk):
        variant = get_object_or_404(Variant, pk=pk)
        name = request.data.get("name", variant.name).strip()
        variant.name = name
        variant.save()
        # Update values if provided
        new_values = request.data.get("values", None)
        if new_values is not None:
            variant.values.all().delete()
            for val in new_values:
                v = val.strip() if isinstance(val, str) else val.get("value", "").strip()
                if v:
                    VariantValue.objects.get_or_create(variant=variant, value=v)
        serializer = VariantSerializer(variant)
        return Response(serializer.data)

    def delete(self, request, pk):
        variant = get_object_or_404(Variant, pk=pk)
        variant.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class VariantValuesView(APIView):

    def post(self, request, pk):
        variant = get_object_or_404(Variant, pk=pk)
        value = request.data.get("value", "").strip()
        if not value:
            return Response({"error": "Value is required"}, status=status.HTTP_400_BAD_REQUEST)
        vv, _ = VariantValue.objects.get_or_create(variant=variant, value=value)
        return Response({"id": vv.id, "value": vv.value}, status=status.HTTP_201_CREATED)


class VariantValueDetailView(APIView):

    def delete(self, request, pk, value_pk):
        vv = get_object_or_404(VariantValue, pk=value_pk, variant_id=pk)
        vv.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class VariantStatsView(APIView):
    def get(self, request):
        total = Variant.objects.count()
        total_values = VariantValue.objects.count()
        return Response({"total": total, "total_values": total_values})

class AllProductsImportView(APIView):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            created_count = 0
            for row in reader:
                name = row.get("name") or row.get("Name")
                sku = row.get("sku") or row.get("SKU")
                cost_price = row.get("cost_price") or row.get("Cost Price") or 0
                sales_price = row.get("sales_price") or row.get("Sales Price") or 0
                
                if name and sku:
                    Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'name': name,
                            'cost_price': cost_price,
                            'sales_price': sales_price,
                            'stock_quantity': row.get("stock_quantity") or row.get("Stock") or 0,
                        }
                    )
                    created_count += 1
            
            return Response({"message": f"Successfully imported {created_count} products."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
